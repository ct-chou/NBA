import string
import requests
import pandas as pd
import openpyxl
import datetime
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl.styles import Alignment, Font, PatternFill


year = 2021  # input("Which NBA season are you interested in?: ")
# input("For which player do you want to get stats?: ")
player_review = ['Giannis Antetokounmpo', 'Khris Middleton',
                 'Jrue Holiday', 'Devin Booker', 'Chris Paul', 'Deandre Ayton']
    #, 'Jae Crowder', 'Pat Connaughton', 'Brook Lopez', 'Bobby Portis']
#player_review = ['Chris Paul', 'Brook Lopez', 'Devin Booker']

#####  1) extract urls for each player #################
url = 'https://www.basketball-reference.com/leagues/NBA_{}_per_game.html'.format(
    year)
r = requests.get(url)
r_html = r.text
soup = BeautifulSoup(r_html, 'html.parser')
table = soup.find_all(class_="full_table")

""" Extracting List of column names"""
head = soup.find(class_="thead")
column_names_raw = [head.text for item in head][0]
column_names_polished = column_names_raw.replace("\n", ",").split(",")[
    2:-1]
total_games = []
"""Extracting full list of player_data"""
players = []

for i in range(len(table)):
    td = table[i].find("td")
    player_name = td.text
    player_url = td.a['href']
    players.append([player_name, player_url])
df = pd.DataFrame(
    players, columns=['Player', 'url']).set_index("Player")
# cleaning the player's name from occasional special characters
df.index = df.index.str.replace('*', '')

df_player_review = df.loc[player_review]
################################################

# 2) loop through player list and download stats for each

dfPlayoff_games_all = []

for i in range(len(df_player_review.index)):
    url = 'https://www.basketball-reference.com{}/gamelog/{}'.format(
        df_player_review['url'][i][:-5], year)
    # grab playoff data & export to excel
    browser = webdriver.Chrome('C:/Users/ctcho/sel_win/chromedriver')
    browser.get(url)
    soup = BeautifulSoup(browser.page_source, 'html.parser')
    table = soup.find('div', {'id': "div_pgl_basic_playoffs"})
    head = table.find('thead') # soup.find() allows us to grab this
    column_names_raw = [head.text for item in head][0]
    column_names_polished = column_names_raw.replace("\n", ",").split(",")[3:-2]
    column_names_polished[4] = 'home'
    column_names_polished[6] = 'W/L'

    body = table.find('tbody')
    body = body.find_all('tr')
    playoff_games = []
    for i in range(len(body)):
        game_stats_ = []
        for td in body[i].find_all("td"):
            game_stats_.append(td.text)
        playoff_games.append(game_stats_)

    dfPlayoff_games = pd.DataFrame(playoff_games, columns=column_names_polished).set_index('G')
    dfPlayoff_games = dfPlayoff_games.dropna()
    dfPlayoff_games['home'] = dfPlayoff_games['home'].mask(dfPlayoff_games['home']=='@','away')
    dfPlayoff_games['home'] = dfPlayoff_games['home'].mask(dfPlayoff_games['home']=='','home')

    dfPlayoff_games_win_loss = dfPlayoff_games['W/L'].astype(str).str[0]
    dfPlayoff_games_result_margin = dfPlayoff_games['W/L'].astype(str).str[1:].replace('(','').replace(')','')

    dfPlayoff_games['W/L'] = dfPlayoff_games_win_loss
    dfPlayoff_games_result_margin = list(map(lambda x: int(x.replace('(','').replace(')','')), dfPlayoff_games_result_margin))
    dfPlayoff_games.insert(6,'W/L Diff', dfPlayoff_games_result_margin)
    dfPlayoff_games['MP'] = list(map(lambda x: int(x.split(':')[0]) + int(x.split(':')[1])/60, dfPlayoff_games['MP']))
    dfPlayoff_games['GS'] = dfPlayoff_games['GS'].astype(bool)
    dfPlayoff_games = dfPlayoff_games.drop(columns=['Age', 'Tm', 'GS','ORB','DRB','GmSc','+/-'])
    total_games.append(len(dfPlayoff_games))
    dfPlayoff_games_all.append(dfPlayoff_games)
    browser.close()

today = datetime.date.today()

workbook_name = './stats/Playoff_stats ' + '{:%y-%m-%d}'.format(today) + '.xlsx'
dfSummary = pd.DataFrame([dfPlayoff_games_all[0].median()], index = [player_review[0]])
dfSeries = pd.DataFrame([dfPlayoff_games_all[0].median()], index = [player_review[0]])
dfHome = pd.DataFrame([dfPlayoff_games_all[0].median()], index = [player_review[0]])
dfAway = pd.DataFrame([dfPlayoff_games_all[0].median()], index = [player_review[0]])
dfWins = pd.DataFrame([dfPlayoff_games_all[0].median()], index = [player_review[0]])
dfLosses = pd.DataFrame([dfPlayoff_games_all[0].median()], index = [player_review[0]])

dfAll5 = pd.DataFrame([dfPlayoff_games_all[0].median()], index = [player_review[0]])
dfSeries5 = pd.DataFrame([dfPlayoff_games_all[0].median()], index = [player_review[0]])
dfAway5 = pd.DataFrame([dfPlayoff_games_all[0].median()], index = [player_review[0]])
dfHome5 = pd.DataFrame([dfPlayoff_games_all[0].median()], index = [player_review[0]])
dfWins5 = pd.DataFrame([dfPlayoff_games_all[0].median()], index = [player_review[0]])
dfLosses5 = pd.DataFrame([dfPlayoff_games_all[0].median()], index = [player_review[0]])


with pd.ExcelWriter(workbook_name) as writer:
    for i in range(len(dfPlayoff_games_all)):
        dfPlayoff_games_all[i] = dfPlayoff_games_all[i].apply(pd.to_numeric, errors='ignore')
        summary_ = pd.DataFrame([dfPlayoff_games_all[i].median()], index = [player_review[i]])
        dfSummary = dfSummary.append(summary_)
        dfPlayoff_games_all[i].to_excel(writer, sheet_name=player_review[i])
        start_row = len(dfPlayoff_games_all[i])
        series_opp = dfPlayoff_games_all[i].iloc[start_row-1]['Opp']
        dfSeries_games = dfPlayoff_games_all[i][dfPlayoff_games_all[i]['Opp'] == series_opp]
        series_ = pd.DataFrame([dfSeries_games.median()], index=[player_review[i]])
        dfSeries = dfSeries.append(series_)
#        dfSeries_games.to_excel(writer, startrow=start_row + 2, sheet_name=player_review[i])
 #       start_row += 4 + len(dfPlayoff_games_all[i][dfPlayoff_games_all[i]['Opp'] == series_opp])
        dfHome_games = dfPlayoff_games_all[i][dfPlayoff_games_all[i]['home'] == 'home']
        home_ = pd.DataFrame([dfHome_games.median()], index=[player_review[i]])
        dfHome = dfHome.append(home_)
        #      dfHome_games.to_excel(writer, startrow=start_row, sheet_name=player_review[i])
   #     start_row += 2 + len(dfPlayoff_games_all[i][dfPlayoff_games_all[i]['home'] == 'home'])
        dfAway_games = dfPlayoff_games_all[i][dfPlayoff_games_all[i]['home'] == 'away']
        away_ = pd.DataFrame([dfAway_games.median()], index=[player_review[i]])
        dfAway = dfAway.append(away_)
    #    dfAway_games.to_excel(writer, startrow=start_row, sheet_name=player_review[i])
        dfWin_game = dfPlayoff_games_all[i][dfPlayoff_games_all[i]['W/L'] == 'W']
        wins_ = pd.DataFrame([dfWin_game.median()], index = [player_review[i]])
        dfWins = dfWins.append(wins_)
        dfLost_game = dfPlayoff_games_all[i][dfPlayoff_games_all[i]['W/L'] == 'L']
        loss_ = pd.DataFrame([dfLost_game.median()], index = [player_review[i]])
        dfLosses = dfLosses.append(loss_)

        ####### last 5 games
        # manual: change series5_
        # todo: adjust if there are less than 5 games e.g. series5
        all5_ = pd.DataFrame([dfPlayoff_games_all[i][-5:].median()], index=[player_review[i]])
        dfAll5 = dfAll5.append(all5_)
        series5_ = pd.DataFrame([dfSeries_games[-5:].median()], index=[player_review[i]])
        dfSeries5 = dfSeries5.append(series5_)
        home5_ = pd.DataFrame([dfHome_games[-5:].median()], index=[player_review[i]])
        dfHome5 = dfHome5.append(home5_)
        away5_ = pd.DataFrame([dfAway_games[-5:].median()], index=[player_review[i]])
        dfAway5 = dfAway5.append(away5_)
        wins5_ = pd.DataFrame([dfWin_game[-5:].median()], index=[player_review[i]])
        dfWins5 = dfWins5.append(wins5_)
        loss5_ = pd.DataFrame([dfLost_game[-5:].median()], index=[player_review[i]])
        dfLosses5 = dfLosses5.append(loss5_)

    dfSummary = dfSummary.iloc[1:]
    dfSeries = dfSeries.iloc[1:]
    dfHome = dfHome.iloc[1:]
    dfAway = dfAway.iloc[1:]
    dfWins = dfWins.iloc[1:]
    dfLosses = dfLosses.iloc[1:]

    dfSummary = dfSummary[['PTS', 'AST', 'TRB', '3P', '3PA', 'MP', 'FG', 'FGA', 'FT', 'FTA', 'TOV', 'PF']]
    dfSeries = dfSeries[['PTS', 'AST', 'TRB', '3P', '3PA', 'MP', 'FG', 'FGA', 'FT', 'FTA', 'TOV', 'PF']]
    dfHome = dfHome[['PTS', 'AST', 'TRB', '3P', '3PA', 'MP', 'FG', 'FGA', 'FT', 'FTA', 'TOV', 'PF']]
    dfAway = dfAway[['PTS', 'AST', 'TRB', '3P', '3PA', 'MP', 'FG', 'FGA', 'FT', 'FTA', 'TOV', 'PF']]
    dfWins = dfWins[['PTS', 'AST', 'TRB', '3P', '3PA', 'MP', 'FG', 'FGA', 'FT', 'FTA', 'TOV', 'PF']]
    dfLosses = dfLosses[['PTS', 'AST', 'TRB', '3P', '3PA', 'MP', 'FG', 'FGA', 'FT', 'FTA', 'TOV', 'PF']]

    dfSeries.to_excel(writer, startrow=1, sheet_name='Median')
    total_players = len(player_review)
    start_row = total_players + 4
    dfSummary.to_excel(writer, startrow=start_row, sheet_name='Median')
    start_row += total_players + 3
    dfHome.to_excel(writer, startrow=start_row, sheet_name='Median')
    start_row += total_players + 3
    dfAway.to_excel(writer, startrow=start_row, sheet_name='Median')
    start_row += total_players + 3
    dfWins.to_excel(writer, startrow=start_row, sheet_name='Median')
    start_row += total_players + 3
    dfLosses.to_excel(writer, startrow=start_row, sheet_name='Median')

# Write Median for last 5 games
    dfAll5 = dfAll5.iloc[1:]
    dfSeries5 = dfSeries5.iloc[1:]
    dfHome5 = dfHome5.iloc[1:]
    dfAway5 = dfAway5.iloc[1:]
    dfWins5 = dfWins5.iloc[1:]
    dfLosses5 = dfLosses5.iloc[1:]

    dfAll5 = dfAll5[['PTS', 'AST', 'TRB', '3P', '3PA', 'MP', 'FG', 'FGA', 'FT', 'FTA', 'TOV', 'PF']]
    dfSeries5 = dfSeries5[['PTS', 'AST', 'TRB', '3P', '3PA', 'MP', 'FG', 'FGA', 'FT', 'FTA', 'TOV', 'PF']]
    dfHome5 = dfHome5[['PTS', 'AST', 'TRB', '3P', '3PA', 'MP', 'FG', 'FGA', 'FT', 'FTA', 'TOV', 'PF']]
    dfAway5 = dfAway5[['PTS', 'AST', 'TRB', '3P', '3PA', 'MP', 'FG', 'FGA', 'FT', 'FTA', 'TOV', 'PF']]
    dfWins5 = dfWins5[['PTS', 'AST', 'TRB', '3P', '3PA', 'MP', 'FG', 'FGA', 'FT', 'FTA', 'TOV', 'PF']]
    dfLosses5 = dfLosses5[['PTS', 'AST', 'TRB', '3P', '3PA', 'MP', 'FG', 'FGA', 'FT', 'FTA', 'TOV', 'PF']]

    start_row = 1
    dfSeries5.to_excel(writer, startrow=start_row, sheet_name='Median Last 5G')
    total_players = len(player_review)
    start_row = total_players + 4
    dfAll5.to_excel(writer, startrow=start_row, sheet_name='Median Last 5G')
    start_row += total_players + 3
    dfHome5.to_excel(writer, startrow=start_row, sheet_name='Median Last 5G')
    start_row += total_players + 3
    dfAway5.to_excel(writer, startrow=start_row, sheet_name='Median Last 5G')
    start_row += total_players + 3
    dfWins5.to_excel(writer, startrow=start_row, sheet_name='Median Last 5G')
    start_row += total_players + 3
    dfLosses5.to_excel(writer, startrow=start_row, sheet_name='Median Last 5G')


# format excel file with openpyxl
wb = openpyxl.load_workbook(workbook_name)

blue = openpyxl.styles.colors.Color(rgb='5FA2FA')
light_blue = openpyxl.styles.colors.Color(rgb='E3F8FC')
fill_blue = PatternFill('solid', fgColor=blue)
fill_light_blue = PatternFill('solid', fgColor=light_blue)

######## format Median Tab ##########
ws = wb['Median']
ws.sheet_view.showGridLines = False

total_columns = len(dfSummary.columns) #start from col 2
start_row = 1

ws['A1'] = 'Current Series'
ws['A1'].font = Font(bold=True,size=18,color='00FF9900')
for column in list(string.ascii_uppercase[1:total_columns+1]):
    for row in range(start_row+1, start_row + 1 + total_players):
        if row % 2 == 0:
            ws[column + str(row + 1)].fill = fill_light_blue

start_row = total_players + 4
ws['A' + str(start_row)] = '2021 Playoffs'
ws['A' + str(start_row)].font = Font(bold=True,size=18,color='00FF9900')
for column in list(string.ascii_uppercase[1:total_columns+1]):
    for row in range(start_row+1, start_row + 1 + total_players):
        if row % 2 == 0:
            ws[column + str(row + 1)].fill = fill_light_blue

start_row += total_players + 3
ws['A' + str(start_row)] = 'Home Games'
ws['A' + str(start_row)].font = Font(bold=True,size=18,color='00FF9900')
for column in list(string.ascii_uppercase[1:total_columns+1]):
    for row in range(start_row+1, start_row + 1 + total_players):
        if row % 2 == 0:
            ws[column + str(row + 1)].fill = fill_light_blue

start_row += total_players + 3
ws['A' + str(start_row)] = 'Away Games'
ws['A' + str(start_row)].font = Font(bold=True,size=18,color='00FF9900')
for column in list(string.ascii_uppercase[1:total_columns+1]):
    for row in range(start_row+1, start_row + 1 + total_players):
        if row % 2 == 0:
            ws[column + str(row + 1)].fill = fill_light_blue

start_row += total_players + 3
ws['A' + str(start_row)] = 'In Games Won'
ws['A' + str(start_row)].font = Font(bold=True,size=18,color='00FF9900')
for column in list(string.ascii_uppercase[1:total_columns+1]):
    for row in range(start_row+1, start_row + 1 + total_players):
        if row % 2 == 0:
            ws[column + str(row + 1)].fill = fill_light_blue

start_row += total_players + 3
ws['A' + str(start_row)] = 'In Games Lost'
ws['A' + str(start_row)].font = Font(bold=True,size=18,color='00FF9900')
for column in list(string.ascii_uppercase[1:total_columns+1]):
    for row in range(start_row+1, start_row + 1 + total_players):
        if row % 2 == 0:
            ws[column + str(row + 1)].fill = fill_light_blue

###### Format Median last 5 games tab ####
ws = wb['Median Last 5G']
ws.sheet_view.showGridLines = False

total_columns = len(dfSummary.columns) #start from col 2
start_row = 1

ws['A1'] = 'Current Series'
ws['A1'].font = Font(bold=True,size=18,color='00FF9900')
for column in list(string.ascii_uppercase[1:total_columns+1]):
    for row in range(start_row+1, start_row + 1 + total_players):
        if row % 2 == 0:
            ws[column + str(row + 1)].fill = fill_light_blue

start_row = total_players + 4
ws['A' + str(start_row)] = '2021 Playoffs'
ws['A' + str(start_row)].font = Font(bold=True,size=18,color='00FF9900')
for column in list(string.ascii_uppercase[1:total_columns+1]):
    for row in range(start_row+1, start_row + 1 + total_players):
        if row % 2 == 0:
            ws[column + str(row + 1)].fill = fill_light_blue

start_row += total_players + 3
ws['A' + str(start_row)] = 'Home Games'
ws['A' + str(start_row)].font = Font(bold=True,size=18,color='00FF9900')
for column in list(string.ascii_uppercase[1:total_columns+1]):
    for row in range(start_row+1, start_row + 1 + total_players):
        if row % 2 == 0:
            ws[column + str(row + 1)].fill = fill_light_blue

start_row += total_players + 3
ws['A' + str(start_row)] = 'Away Games'
ws['A' + str(start_row)].font = Font(bold=True,size=18,color='00FF9900')
for column in list(string.ascii_uppercase[1:total_columns+1]):
    for row in range(start_row+1, start_row + 1 + total_players):
        if row % 2 == 0:
            ws[column + str(row + 1)].fill = fill_light_blue

start_row += total_players + 3
ws['A' + str(start_row)] = 'In Games Won'
ws['A' + str(start_row)].font = Font(bold=True,size=18,color='00FF9900')
for column in list(string.ascii_uppercase[1:total_columns+1]):
    for row in range(start_row+1, start_row + 1 + total_players):
        if row % 2 == 0:
            ws[column + str(row + 1)].fill = fill_light_blue

start_row += total_players + 3
ws['A' + str(start_row)] = 'In Games Lost'
ws['A' + str(start_row)].font = Font(bold=True,size=18,color='00FF9900')
for column in list(string.ascii_uppercase[1:total_columns+1]):
    for row in range(start_row+1, start_row + 1 + total_players):
        if row % 2 == 0:
            ws[column + str(row + 1)].fill = fill_light_blue


#format individual player tabs
for player in range(len(player_review)):
    ws = wb[player_review[player]]
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 9.8
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 4
    ws.column_dimensions['F'].width = 6.4

    for column in ('H', 'I', 'J', 'K','L','M','N','O','P','Q','R','S','T','U','V','W'):
        ws.column_dimensions[column].width = 5.5

    for column in ('A','B','C','D','E','F','G','H', 'I', 'J', 'K','L','M','N','O','P','Q','R','S','T','U','V','W'):
        for j in range(total_games[player]+1):
            if (j+1) % 2 == 0:
                ws[column+str(j+1)].fill = fill_light_blue

wb.save(workbook_name)

# filter one by wins and one by losses export to excel


#
    # filter by home games and away games


